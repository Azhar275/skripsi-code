from ultralytics import YOLO
import random
from deepsort import Tracker
import math
from threading import Timer
import time
from multiprocessing import Process
import pandas as pd

# importing the Computer Vision module
import cv2
from flask import Flask,Response,render_template,request,redirect,session,url_for
import mysql.connector
import numpy as np
from flask_socketio import SocketIO, emit

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy_utils import create_database, database_exists
# from sqlalchemy.sql import row_number #import sqlalchemy functions
from sqlalchemy import text, func, distinct, desc, update,create_engine,select, insert, tuple_ #import sqlalchemy update function and create_engine function
# from sqlalchemy.sql.expression import row_number
from sqlalchemy.orm import aliased, sessionmaker #import sqlalchemy session maker function
import pymysql

# login library
# from flask_login import UserMixin
from flask_bcrypt import Bcrypt
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SelectField
from wtforms.validators import InputRequired
from flask_bootstrap import Bootstrap
from functools import wraps

#app scheduler
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
import calendar
# import torch
# torch.cuda.set_device(0)

# xlwt library for excel export
# import xlwt
import io
from babel.dates import format_date, Locale
locale = Locale('id_ID')
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, Alignment, Border, Side, PatternFill


# import torch
# torch.cuda.set_device(0)

app=Flask(__name__,static_url_path='/static') #initializing the flask app with the name 'app' and static_url_path for static files
socketio=SocketIO(app) #socketio initialization for real-time communication
app.app_context().push()
# ========== DB CONNECTION -- SQLALCHEMY (START) ===========
#Database Configuration with MySQL
#     -->     db type | username: password | path to database name    
# engine_url='mysql+pymysql://root:''@localhost/enpemo'
#     Server Configuration with MySQL  
engine_url='mysql+pymysql://root:''@localhost/enpemo'

if not database_exists(engine_url):
    create_database(engine_url)
app.config['SQLALCHEMY_DATABASE_URI']=engine_url
app.config['SECRET_KEY'] = "enpemoaja123"
engine=create_engine(engine_url)
Session=sessionmaker(engine)
db=SQLAlchemy(app) #initializing the database with the name 'db'
class PolygonCoordinates(db.Model):
    id=db.Column(db.Integer,primary_key=True,autoincrement=True)
    preference_num=db.Column(db.Integer,nullable=False)
    x1=db.Column(db.Integer,nullable=False)
    y1=db.Column(db.Integer,nullable=False)
    x2=db.Column(db.Integer,nullable=False)
    y2=db.Column(db.Integer,nullable=False)
    x3=db.Column(db.Integer,nullable=False)
    y3=db.Column(db.Integer,nullable=False)
    x4=db.Column(db.Integer,nullable=False)
    y4=db.Column(db.Integer,nullable=False)
    createdAt=db.Column(db.DateTime,nullable=False)
    updatedAt=db.Column(db.DateTime,nullable=False)

    def __init__(self,preference_num,x1,y1,x2,y2,x3,y3,x4,y4,createdAt,updatedAt):
        self.preference_num=preference_num
        self.x1=x1
        self.y1=y1
        self.x2=x2
        self.y2=y2
        self.x3=x3
        self.y3=y3
        self.x4=x4
        self.y4=y4
        self.createdAt=createdAt
        self.updatedAt=updatedAt

class Occupancy(db.Model):
    __tablename__ = 'd_occupancy'
    id=db.Column(db.Integer,primary_key=True,autoincrement=True)
    enter_total=db.Column(db.Integer,nullable=False)
    out_total=db.Column(db.Integer,nullable=False)
    in_room=db.Column(db.Integer,nullable=False)
    createdAt=db.Column(db.DateTime,nullable=False)

    def __init__(self, enter_total, out_total, in_room, createdAt):
        self.enter_total = enter_total
        self.out_total = out_total
        self.in_room = in_room
        self.createdAt = createdAt

class Average(db.Model):
    __tablename__ = 'd_average'
    id=db.Column(db.Integer,primary_key=True,autoincrement=True)
    enter_total=db.Column(db.Integer,nullable=False)
    out_total=db.Column(db.Integer,nullable=False)
    in_room=db.Column(db.Integer,nullable=False)
    createdAt=db.Column(db.DateTime,nullable=False)

    def __init__(self, enter_total, out_total, in_room, createdAt):
        self.enter_total = enter_total
        self.out_total = out_total
        self.in_room = in_room
        self.createdAt = createdAt
# ========== DB CONNECTION -- SQLALCHEMY (END) ===========


# ========== GETTER AND SETTER [COORDINATES] FUNCTION (START) ===========
'''Getting Coordinate from The Database'''

def getCoordinates():
    poly_coordinates={} #default value for polygon coordinates 
    with app.app_context(): # create a context for the database access
        # query for getting the last row of coordinates
        # Using ORM Approach
        result=PolygonCoordinates.query.filter_by(id=1).first()
        if result:
            #assigning the result to the global variable
            poly_coordinates['x1'] = result.x1
            poly_coordinates['y1'] = result.y1
            poly_coordinates['x2'] = result.x2
            poly_coordinates['y2'] = result.y2
            poly_coordinates['x3'] = result.x3
            poly_coordinates['y3'] = result.y3
            poly_coordinates['x4'] = result.x4
            poly_coordinates['y4'] = result.y4
        else:
            newPolyCoordinates = PolygonCoordinates(1,200,300,500,300,500,100,200,100,func.now(),func.now())
            db.session.add(newPolyCoordinates)
            db.session.commit()
            print("No data found, inserting default data...")
    return poly_coordinates


@app.route('/submit_coordinates',methods=['POST'])
def submitCoordinates():
    # get the hidden input form value from the html templates
    coordinates=request.form.get('coordinates')
    print(coordinates)
    # split the coordinates value into array
    coordinates=coordinates.split(' ')
    print ("received coordinates -> ",coordinates)
    # query for updating the coordinates value
    # ORM Approach for Update Coordinate
    updated_coordinates = {
        'x1': coordinates[0],
        'y1': coordinates[1],
        'x2': coordinates[2],
        'y2': coordinates[3],
        'x3': coordinates[4],
        'y3': coordinates[5],
        'x4': coordinates[6],
        'y4': coordinates[7]
    }
    with app.app_context(): # create a context for the database access
        PolygonCoordinates.query.filter_by(id=1).update(updated_coordinates)
        db.session.commit()
  
    return redirect('/settings')

def submitData():
    with app.app_context():
        new_data = Occupancy(len(enter_list), len(out_list), len(enter_list)-len(out_list), func.now())
        db.session.add(new_data)
        db.session.commit()
        print(f"STORE TO DATABASE VALUE: {len(enter_list)} {len(out_list)} {len(enter_list)-len(out_list)} {func.now()}")
        return True
# ========== GETTER AND SETTER [COORDINATES] FUNCTION (END) ===========

# ========== VIDEO STREAM (START) ===========

# cap = cv2.VideoCapture(0) #using webcam
# cap = cv2.VideoCapture('rtsp://admin:admin123@192.168.22.176:554/Streaming/Channels/201') #using ip camera
# cap = cv2.VideoCapture(0) #using ip camera
tracker=Tracker()

people_list = {}
enter_list = {}
out_list = {}

colors = [(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255)) for j in range(10)]

def distance(x1, y1, x2, y2):
  num_a = math.pow((x2-x1), 2)
  num_b = math.pow((y2-y1), 2)
  result = math.sqrt(num_a + num_b)
  return result

# Fungsi untuk reset variabel pada pukul 00.00
def reset_at_midnight():
    current_time = datetime.now()
    print(current_time.hour, current_time.minute)
    if current_time.hour == 0 and current_time.minute == 0:
        # Reset semua variabel yang Anda inginkan di sini
        enter_list.clear()
        out_list.clear()
        people_list.clear()

def distance(x1, y1, x2, y2):
  num_a = math.pow((x2-x1), 2)
  num_b = math.pow((y2-y1), 2)
  result = math.sqrt(num_a + num_b)
  return result

def timerImage(cap):
    timer = Timer(0.5, lambda: None)  # just returns None on timeout
    timer.start()
    frame = None
    success = None
    failed = True
    while timer.is_alive():
        success, frame = cap.read()
        failed = False
        return success, frame, failed
    return success, frame, failed

# def capture_image(cap):
#     start_time = time.time()
#     frame = None
#     success = False
#     while time.time() - start_time < 5:
#         success, frame = cap.read()
#         if success:
#             break
#     if not success:
#         print("Camera not found")
#         return False, None
#     else:
#         print("Camera found")
#         return success, frame

def reloadCamera():
    return redirect('/')

# polygon zone preview
def annotatedStream():
    # cap = cv2.VideoCapture("rtsp://admin:admin123@sg2.labkom.us:1736/Streaming/Channels/201") #indoor office cctv
    cap = cv2.VideoCapture("rtsp://admin:TCHIOT@192.168.1.4:554/Streaming/Channels/101")
    # cap = cv2.VideoCapture(0)  
    # temporarily stored here
    poly_zone=getCoordinates()
    polygon_zone=np.array([[poly_zone['x1'],poly_zone['y1']],
                        [poly_zone['x2'],poly_zone['y2']],
                        [poly_zone['x3'],poly_zone['y3']],
                        [poly_zone['x4'],poly_zone['y4']]],
                        np.int32)
    model = YOLO('yolov8n.pt')
    reloaded = 0    
    while True:
        # reset_at_midnight()
        reset_at_midnight()
        # reading frames from the video
        # success, frame = cap.read()
        failed = False
        success, frame, failed = timerImage(cap)
        if success == False:
            print('entered 2')
            cap = cv2.VideoCapture("https://cdn.osxdaily.com/wp-content/uploads/2013/12/there-is-no-connected-camera-mac.jpg")
            success, frame = cap.read()
            frame=cv2.resize(frame,(880,440))
            # buffer=cv2.imencode('.jpg',frame)[1] # change to index 1 to get the buffer
            # frame=buffer.tobytes()# converting the image to bytes
            # yield(b'--frame\r\n' # yielding the frame for display
            #       b'Content-Type: image/jpeg\r\n\r\n'+frame+b'\r\n')
            time.sleep(5)
            # cap = cv2.VideoCapture("rtsp://admin:admin123@sg2.labkom.us:1736/Streaming/Channels/201")
            cap = cv2.VideoCapture("rtsp://admin:admin123@192.168.25.254:554/Streaming/Channels/201")  
            # cap = cv2.VideoCapture(0)
        
        if failed == True :
            print('Reload')
            print('Please')
            reloaded+=1
            # cap = cv2.VideoCapture("rtsp://admin:admin123@sg2.labkom.us:5465/Streaming/Channels/201") #indoor office cctv
            cap = cv2.VideoCapture("rtsp://admin:admin123@192.168.25.254:554/Streaming/Channels/201")  
            # cap = cv2.VideoCapture(0)
            continue
            # break
        else:
            # converting the collected frame to image
            frame=cv2.resize(frame,(880,440))
            results = model.predict(frame, classes=0, stream_buffer=True, stream=True)
            
            print('reloaded ', reloaded)
            for result in results:
                detections = []
                for r in result.boxes.data.tolist():
                    x1, y1, x2, y2, score, class_id = r
                    x1 = int(x1)
                    y1 = int(y1)
                    x2 = int(x2)
                    y2 = int(y2)
                    class_id = int(class_id)
                    detections.append([x1, y1, x2, y2, score])
                tracker.update(frame, detections)
                for track in tracker.tracks:
                    bbox = track.bbox
                    x1, y1, x2, y2 = bbox
                    x1 = int(x1)
                    y1 = int(y1)
                    x2 = int(x2)
                    y2 = int(y2)
                    track_id = track.track_id

                    xa = poly_zone['x1']
                    ya = poly_zone['y1']
                    xb = poly_zone['x2']
                    yb = poly_zone['y2']
                    xc = poly_zone['x3']
                    yc = poly_zone['y3']
                    xd = poly_zone['x4']
                    yd = poly_zone['y4']

                    result_a = distance(xa, ya, x2, y2) + distance(xb, yb, x2, y2)
                    result_b = distance(xa, ya, xb, yb)
                    result1 = round(result_a, 3) - round(result_b, 3)

                    result_c = distance(xc, yc, x2, y2) + distance(xd, yd, x2, y2)
                    result_d = distance(xc, yc, xd, yd)
                    result3 = round(result_c, 3) - round(result_d, 3)

                    # =========================OLD==============================
                    # Line counter
                    # m1 = (poly_zone['y2'] - poly_zone['y1'])/(poly_zone['x2'] - poly_zone['x1'])
                    # b1 = poly_zone['y1'] - (poly_zone['x1']*m1)
                    # result1 = y2 - ((m1*x2)+b1)
                    # print(result1)

                    # m2 = (poly_zone['y4'] - poly_zone['y3'])/(poly_zone['x4'] - poly_zone['x3'])
                    # b2 = poly_zone['y3'] - (poly_zone['x3']*m2)
                    # result2 = y2 - ((m2*x2)+b2)
                    # print(result2)

                    if  result1 <= 1 and result1 >= -1:
                        if track_id not in people_list.keys():
                            people_list[track_id] = 'in'
                        if people_list[track_id] == 'out':
                            enter_list[track_id] = [x2, y2]
                            people_list.pop(track_id)
                    
                    if  result3 <= 1 and result3 >= -1:
                        if track_id not in people_list.keys():
                            people_list[track_id] = 'out'
                        if people_list[track_id] == 'in':
                            out_list[track_id] = [x2, y2]
                            people_list.pop(track_id)

                    # ======================OLD V2=============================

                    # if  result1 <= 1 and result1 >= -1:
                    #     if track_id not in people_list.keys():
                    #         people_list[track_id] = 'in'
                    #     if track_id not in out_list.keys() and people_list[track_id] == 'out':
                    #         enter_list[track_id] = [x2, y2]
                    
                    # if  result3 <= 1 and result3 >= -1:
                    #     if track_id not in people_list.keys():
                    #         people_list[track_id] = 'out'
                    #     if track_id not in enter_list.keys() and people_list[track_id] == 'in':
                    #         out_list[track_id] = [x2, y2]

                    # ===============OLD=========================

                    # if (result1 <= 3 and result1 >= 0) or (result1 >= (-3) and result1 <= 0)  :
                    #     if track_id not in out_list.keys():
                    #         enter_list[track_id] = [x2, y2]

                    # if (result2 <= 3 and result2 >= 0) or (result2 >= (-3) and result2 <= 0)  :
                    #     if track_id not in enter_list.keys():
                    #         out_list[track_id] = [x2, y2]

                    dist = cv2.pointPolygonTest(polygon_zone, (x2,y2), False)
                    # if dist == 1 and (track_id in out_list.keys() or track_id in enter_list.keys()):
                    #     people_list[track_id] = y2
                    cv2.putText(frame, (str(track_id)),(x1,y1),cv2.FONT_HERSHEY_COMPLEX,0.8,(0,0,255),2)
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (colors[track_id % len(colors)]), 3)
            # print('Enter : ', enter_list)
            # print('Out : ', out_list)
            # annotate the frame using polygon
            
            frame=cv2.polylines(frame,[np.array(polygon_zone,np.int32)],True,(0,0,0),4,cv2.LINE_AA)
            cv2.line(frame,(poly_zone['x1'], poly_zone['y1']),(poly_zone['x2'], poly_zone['y2']),(0,255,0),2)
            


            cv2.line(frame,(poly_zone['x3'], poly_zone['y3']),(poly_zone['x4'], poly_zone['y4']),(0,0,255),2)
            cv2.rectangle(frame, (0, 40), (200, 130), (255,255,255), -1)
            cv2.putText(frame, ("Melewati Zona :"+str(len(people_list))),(10,60),cv2.FONT_HERSHEY_DUPLEX,0.5,(0,0,0),2)
            cv2.putText(frame, ("Masuk :"+str(len(enter_list))),(10,80),cv2.FONT_HERSHEY_DUPLEX,0.5,(0,0,0),2)
            cv2.putText(frame, ("Keluar :"+str(len(out_list))),(10,100),cv2.FONT_HERSHEY_DUPLEX,0.5,(0,0,0),2)
            cv2.putText(frame, ("Kepadatan :"+str(len(enter_list)-len(out_list))),(10,120),cv2.FONT_HERSHEY_DUPLEX,0.5,(0,0,0),2)
            buffer=cv2.imencode('.jpg',frame)[1] # change to index 1 to get the buffer
            frame=buffer.tobytes()# converting the image to bytes
            yield(b'--frame\r\n' # yielding the frame for display
                  b'Content-Type: image/jpeg\r\n\r\n'+frame+b'\r\n')
    

bcrypt = Bcrypt(app)
bootstrap = Bootstrap(app)

# ==== DB USER ====
class User(db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), nullable=False, unique=True)
    password = db.Column(db.String(80), nullable=False)

    def __init__(self, username, password):
      self.username = username
      if password != '':
         self.password = bcrypt.generate_password_hash(password).decode('UTF-8')

class Login(FlaskForm):
   username = StringField('', validators=[InputRequired()],render_kw={'autofocus':True, 'placeholder':'Username'})
   password = PasswordField('', validators=[InputRequired()],render_kw={'autofocus':True, 'placeholder':'Password'})

def login_req(f):
   @wraps(f)
   def wrap(*args,**kwargs):
      if 'login' in session:
         return f(*args,**kwargs)
      else:
         return redirect(url_for('login'))
   return wrap

@app.route('/login', methods=['GET','POST'])
def login():
   if session.get('login'):
      return redirect(url_for('analytics'))
   form = Login()
   if form.validate_on_submit():
      user = User.query.filter_by(username=form.username.data).first()
      if user:
         if bcrypt.check_password_hash(user.password, form.password.data):
            session['login'] = True
            session['id'] = user.id
            session['username'] = user.username
            stars = ''
            for i in form.password.data:
                stars += '*'
            session['stars'] = stars
            return redirect(url_for('analytics'))
      pesan = "Username atau Password salah"
      return render_template('login.html',pesan=pesan,form=form)
   return render_template('login.html', form=form)

@app.route('/logout')
@login_req
def logout():
   session.clear()
   return redirect(url_for('login'))

@app.route('/')
@login_req
def root():
    return redirect(url_for('login'))

@app.route('/profile')
@login_req
def profile():
    return render_template('profile.html')

@app.route('/editprofile', methods=['POST','GET'])
@login_req
def editprofile():
    data = User.query.filter_by(id=session.get('id')).first()
    if request.method == 'POST':
        data.username = request.form['new_username']
        if request.form['new_password'] == request.form['conf_new_password']:
            if bcrypt.check_password_hash(data.password, request.form['old_password']):
                data.password = bcrypt.generate_password_hash(request.form['new_password']).decode('UTF-8')

                db.session.add(data)
                db.session.commit()

                session['username'] = data.username
                stars = ''
                for i in request.form['new_password']:
                    stars += '*'
                session['stars'] = stars

                return redirect(url_for('profile'))
            alert = 'Gagal mengubah, password saat ini tidak valid '
            return render_template('profile.html', alert=alert)
        alert = 'Gagal mengubah, password baru tidak sama'
        return render_template('profile.html', alert=alert)

def weekFormat(start, end):
    week = f"{start.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}"
    if start.strftime("%Y") == end.strftime("%Y"):
        week = f"{start.strftime('%d %b')} - {end.strftime('%d %b')} {start.strftime('%Y')}"
        if start.strftime("%b") == end.strftime("%b"):
            week = f"{start.strftime('%d')} - {end.strftime('%d')} {start.strftime('%b %Y')}"
    return week

@app.route('/dashboard', methods=['GET', 'POST'])
@login_req
def dashboard():
    labelOfChart = ''

    dateOfDay = datetime.today().date()
    theDay = format_date(datetime.today(), format='EEEE', locale=locale)

    if request.method == 'POST':
        filter_type = request.form['date_type']
        input_date = request.form['filterdate']
        if filter_type == 'date':
            dateOfDay = datetime.strptime(input_date, '%Y-%m-%d').date()
            theDay = format_date(datetime.strptime(input_date, '%Y-%m-%d').date(), format='EEEE', locale=locale)
            labelOfChart = 'day'
        elif filter_type == 'week':
            year, week = map(int, input_date.split('-W'))
            dateOfDay = datetime.strptime(f'{year}-W{week}-1', "%Y-W%W-%w").date()
            theDay = format_date(datetime.strptime(f'{year}-W{week}-1', "%Y-W%W-%w").date(), format='EEEE', locale=locale)
            labelOfChart = 'week'
        elif filter_type == 'month':
            dateOfDay = datetime.strptime(input_date, "%Y-%m").date()

            if dateOfDay.weekday() in (5, 6):
                numberOfDay = dateOfDay.weekday()
                if numberOfDay == 5:
                    dateOfDay = dateOfDay + timedelta(days=2)
                elif numberOfDay == 6:
                    dateOfDay = dateOfDay + timedelta(days=1)

            theDay = format_date(dateOfDay, format='EEEE', locale=locale)
            labelOfChart = 'month'
    
    theYear = dateOfDay.year
    theMonth = dateOfDay.month
    theWeek = dateOfDay.isocalendar()[1]
    weekOfDay = f"{theYear}-W{theWeek:02}"
    monthOfDay = dateOfDay.strftime("%Y-%m")
    
    # DAILY DATA
    data_theDay = Occupancy.query.filter(Occupancy.createdAt >= dateOfDay).filter(Occupancy.createdAt < dateOfDay + timedelta(days=1)).all()
    occupancy_theDay = [item.in_room for item in data_theDay]
    in_theDay = [item.enter_total for item in data_theDay]
    out_theDay = [item.out_total for item in data_theDay]
    hour_theDay = [item.createdAt.strftime("%H:%M") for item in data_theDay]
    
    # WEEKLY DATA
    firstWeekday = dateOfDay - timedelta(days=dateOfDay.weekday())
    endWeekday = firstWeekday + timedelta(days=4)
    # print("Senin",firstWeekday)
    # print("Jumat",endWeekday)
    
    # avgOccupancy_week = Session().query(
    #                 func.avg(Occupancy.in_room)
    #                 ).filter(
    #                     Occupancy.createdAt >= firstWeekday,
    #                     Occupancy.createdAt < endWeekday + timedelta(days=1),
    #                     ).group_by(func.date(Occupancy.createdAt))
    inOutData_week = Session().query(
                    func.max(Occupancy.enter_total),
                    func.max(Occupancy.out_total)
                    ).filter(
                        func.WEEK(Occupancy.createdAt) == theWeek
                        ).group_by(func.date(Occupancy.createdAt))
    # inOutData_week = Session().query(
    #                 func.max(Occupancy.enter_total),
    #                 func.max(Occupancy.out_total),
    #                 func.max(Occupancy.in_room)
    #                 ).filter(
    #                     Occupancy.createdAt >= firstWeekday,
    #                     Occupancy.createdAt < endWeekday + timedelta(days=1),
    #                     ).group_by(func.date(Occupancy.createdAt))
    # inOutData_week = Session().query(
    #     Occupancy.createdAt.label('date'),
    #     func.max(Occupancy.in_room).label('max_enter_total')
    #     ).filter(
    #         Occupancy.createdAt >= firstWeekday,
    #         Occupancy.createdAt < endWeekday + timedelta(days=1),
    #         ).group_by(func.date(Occupancy.createdAt)).all()
    
    subq_occupancy_week = Session().query(
        func.date(Occupancy.createdAt).label('tanggal'),
        func.max(Occupancy.in_room).label('max_in_room'),
    ).filter(func.week(Occupancy.createdAt) == theWeek).group_by(func.date(Occupancy.createdAt)).subquery()

    occupacyData_week = Session().query(
        Occupancy.createdAt,
        Occupancy.in_room
    ).filter(
        tuple_(func.date(Occupancy.createdAt), Occupancy.in_room)
        .in_(subq_occupancy_week)
    ).group_by(func.date(Occupancy.createdAt)).order_by(func.date(Occupancy.createdAt)).all()

    subq_occupancyMin_week = Session().query(
        func.date(Occupancy.createdAt).label('tanggal'),
        func.min(Occupancy.in_room).label('min_in_room'),
    ).filter(func.week(Occupancy.createdAt) == theWeek).group_by(func.date(Occupancy.createdAt)).subquery()

    occupacyDataMin_week = Session().query(
        Occupancy.createdAt,
        Occupancy.in_room
    ).filter(
        tuple_(func.date(Occupancy.createdAt), Occupancy.in_room)
        .in_(subq_occupancyMin_week)
    ).group_by(func.date(Occupancy.createdAt)).order_by(func.date(Occupancy.createdAt)).all()
    
    in_week = [item[0] for item in inOutData_week]
    out_week = [item[1] for item in inOutData_week]
    occupancy_week = [item.in_room for item in occupacyData_week]
    time_week = [f"{item.createdAt.strftime('%H:%M')} WIB" for item in occupacyData_week]
    occupancyMin_week = [item.in_room for item in occupacyDataMin_week]
    timeOccMin_week = [f"{item.createdAt.strftime('%H:%M')} WIB" for item in occupacyDataMin_week]
    # print(time_week)
    # print(occupancyMin_week)
    # print(timeOccMin_week)


    # Iterate and print data
    # for data in occupacyData_week:
    #     print(f"Tanggal: {data.createdAt}, In Room: {data.in_room}")
    
    # MONTHLY DATA
    firstDateOfMonth = dateOfDay.replace(day=1)
    lastDateOfMonth = firstDateOfMonth + timedelta(days=calendar.monthrange(theYear, theMonth)[1]-1)

    firstDateInFirstWeek = firstDateOfMonth - timedelta(days=firstDateOfMonth.weekday())
    lastDateInLastWeek = lastDateOfMonth - timedelta(days=lastDateOfMonth.weekday()) + timedelta(days=6)

    # subq_dailyData_month = Session().query(
    # func.max(Occupancy.enter_total).label("max_enter"),
    # func.max(Occupancy.out_total).label("max_out"),
    # func.max(Occupancy.in_room).label("avg_occupancy"),
    # Occupancy.createdAt.label("week_number")
    # ).filter(
    # Occupancy.createdAt >= firstDateInFirstWeek,
    # Occupancy.createdAt < lastDateInLastWeek + timedelta(days=1),
    # ).group_by(func.date(Occupancy.createdAt)).subquery()


    # data_month = Session().query(
    #                 func.max(subq_dailyData_month.c.max_enter),
    #                 func.max(subq_dailyData_month.c.max_out),
    #                 func.max(subq_dailyData_month.c.avg_occupancy)
    #                 ).group_by(func.extract('week', subq_dailyData_month.c.week_number))

    data = Session().query(
        Occupancy.createdAt.label('date'),
        func.max(Occupancy.in_room)
    ).filter(text('''createdAt in(
        SELECT createdAt
        FROM d_occupancy
        WHERE (in_room, WEEK(createdAt)) IN (
            SELECT MAX(in_room), WEEK(createdAt) FROM d_occupancy GROUP BY WEEK(createdAt)
        )
    )'''),Occupancy.createdAt >= firstDateInFirstWeek,
            Occupancy.createdAt < lastDateInLastWeek + timedelta(days=1)).group_by(func.date(Occupancy.createdAt))

    occupancyMax_month = []
    timeOfOccMax_month = []
    for month_time in data:
        if len(occupancyMax_month) > 4:
            break
        theDayOfTime = format_date(month_time.date, format='EEEE', locale=locale)
        dateOfTheDay = month_time.date.strftime('%d-%m-%y')
        timeIn_theDayOfTime = month_time.date.strftime('%H:%M')
        timeOfOccMax_month.append(f"{theDayOfTime}, {dateOfTheDay}\n{timeIn_theDayOfTime} WIB")
        occupancyMax_month.append(month_time[1])
    # timeOfOccMax_month = [f"{item.date.strftime('%H:%M')} WIB" for item in data]
    # print(timeOfOccMax_month)
    # print(occupancyMax_month)

    data_min = Session().query(
        Occupancy.createdAt.label('date'),
        func.min(Occupancy.in_room)
    ).filter(text('''createdAt in(
        SELECT createdAt
        FROM d_occupancy
        WHERE (in_room, WEEK(createdAt)) IN (
            SELECT MIN(in_room), WEEK(createdAt) FROM d_occupancy GROUP BY WEEK(createdAt)
        )
    )'''),Occupancy.createdAt >= firstDateInFirstWeek,
            Occupancy.createdAt < lastDateInLastWeek + timedelta(days=1)).group_by(func.date(Occupancy.createdAt))

    occupancyMin_month = []
    timeOfOccMin_month = []
    for month_time in data_min:
        if len(occupancyMin_month) > 4:
            break
        theDayOfTime = format_date(month_time.date, format='EEEE', locale=locale)
        dateOfTheDay = month_time.date.strftime('%d-%m-%y')
        timeIn_theDayOfTime = month_time.date.strftime('%H:%M')
        timeOfOccMin_month.append(f"{theDayOfTime}, {dateOfTheDay}\n{timeIn_theDayOfTime} WIB")
        occupancyMin_month.append(month_time[1])
    # timeOfOccMin_month = [f"{item.date.strftime('%H:%M')} WIB" for item in data]
    # print(timeOfOccMin_month)
    # print(occupancyMin_month)
    
    data_month = Session().query(
        Occupancy.createdAt,
        func.max(Occupancy.enter_total),
        func.max(Occupancy.out_total),
        func.max(Occupancy.in_room)
    ).filter(text('''DATE(createdAt) in(
        SELECT DATE(createdAt)
        FROM d_occupancy
        WHERE (in_room, WEEK(createdAt)) IN (
            SELECT MAX(in_room), WEEK(createdAt) FROM d_occupancy GROUP BY WEEK(createdAt)
        )
    )'''),Occupancy.createdAt >= firstDateInFirstWeek,
            Occupancy.createdAt < lastDateInLastWeek + timedelta(days=1)).group_by(func.date(Occupancy.createdAt))
    
    in_month = [data_in[1] for data_in in data_month]
    out_month = [data_out[2] for data_out in data_month]
    # occupancy_month = [data_occupancy[3] for data_occupancy in data_month]
    week_month = [f"Minggu {week}" for week in range(1,len(occupancyMax_month)+1)]

    Session.close_all()
    
    return render_template('index.html', 
                           hari=theDay, today=dateOfDay, dayPeriod=dateOfDay.strftime("%d %B %Y"), weekPeriod=weekFormat(firstWeekday,endWeekday), monthPeriod=dateOfDay.strftime("%B %Y"), weekOfDay=weekOfDay, monthOfDay=monthOfDay, 
                           ruang=occupancy_theDay, masuk=in_theDay, keluar=out_theDay, jam=hour_theDay, 
                           mingguanruang=occupancy_week, mingguankeluar=out_week, mingguanmasuk=in_week, timeWeek=time_week, occupancyMin_week=occupancyMin_week, timeOccMin_week=timeOccMin_week,
                           in_month=in_month, out_month=out_month, occupancyMin_month=occupancyMin_month, occupancyMax_month=occupancyMax_month, timeOfOccMin_month=timeOfOccMin_month, timeOfOccMax_month=timeOfOccMax_month, week_month=week_month, 
                           label=labelOfChart)

@app.route('/download_data',methods=['GET', 'POST'])
@login_req
def download_data():
    if request.method == 'POST':
        str_date_start=request.form['date_start']
        str_date_end=request.form['date_end']
        date_start=datetime.strptime(str_date_start, '%Y-%m-%d')
        date_end=datetime.strptime(str_date_end, '%Y-%m-%d')
        date_end=date_end+timedelta(days=1)
        data = Occupancy.query.filter(Occupancy.createdAt.between(date_start, date_end)).all()
    output=io.BytesIO()
    #create workbook object
    workbook=Workbook()
    # Initial Workbook Configuration - Sheet Name, Column Width, Merge Cell, Font Style
    worksheet=workbook.active
    worksheet.title="Laporan Kepadatan Area"
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 7
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 20
    
    worksheet.merge_cells('A1:H1')
    font_title=Font(name='Calibri',size=14,bold=True)
    title_style=worksheet['A1']
    title_style.font=font_title

    worksheet['A1'] = f"Laporan Jumlah Kepadatan Area Pada {datetime.strptime(str_date_start, '%Y-%m-%d').strftime('%d %B %Y')} hingga {datetime.strptime(str_date_end, '%Y-%m-%d').strftime('%d %B %Y')}"
    worksheet['A2']='' #break
    # Header Configuration
    header_contents=['Tanggal','Jam','Jumlah Masuk','Jumlah Keluar','Kepadatan Ruangan','Rerata Kepadatan']
    worksheet.append(header_contents)
    for col_num in range(1, len(header_contents) + 1):
        font_header=Font(name='Calibri',size=12,bold=True)
        cell = worksheet.cell(row=3, column=col_num)
        cell.font=font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
   
    # initiation of merging cells logic
    previous_date=None
    merge_start=4
    merge_end=0
    current_row=4
    record_state=False
    occupancy=0
    # iterate every row found
    for row in data:
        #if there's no previous date, set the previous date to the current date
        if previous_date==None:
            previous_date=row.createdAt.strftime("%d %B %Y")
            occupancy+=row.in_room
        #to check if the current date is the same as the previous date and the record state is false
        elif previous_date==row.createdAt.strftime("%d %B %Y") and record_state==False:
            #set the merge start and end to the current row
            merge_start=current_row-1
            merge_end=current_row
            record_state=True
            occupancy+=row.in_room
        # other branch: if the current date is the same as the previous date and the record state is true    
        elif previous_date==row.createdAt.strftime("%d %B %Y") and record_state==True:
            #will continue to set the merge end to the current row
            merge_end=current_row
            occupancy+=row.in_room
        # to represent that the date has changed and the record state is true
        elif previous_date!=row.createdAt.strftime("%d %B %Y") and record_state==True:
            # will merge the cells from the merge start to the merge end also set the record state to false and set the previous date to the current date
            worksheet.merge_cells(f'A{merge_start}:A{merge_end}')
            worksheet.merge_cells(f'F{merge_start}:F{merge_end}')
            # for row_num in range(merge_start, merge_end + 1):
            #     cell = worksheet.cell(row=row_num, column=1)
            #     cell.alignment = Alignment(wrap_text=True)
            record_state=False
            previous_date=row.createdAt.strftime("%d %B %Y")
            worksheet[f'F{merge_start}']=round(occupancy/(merge_end-merge_start+1),2)
            occupancy=0
        # Add New Row Data    
        day=format_date(row.createdAt, format='EEEE', locale=locale)
        date=row.createdAt.strftime("%d %B %Y")
        day_date=f"{day}, {date}"
        row_data=[day_date,row.createdAt.strftime("%H:%M"),row.enter_total,row.out_total,row.in_room]
        worksheet.append(row_data)

        # Apply center alignment to each cell in the current row
        # len(row_data) + 2 -> to include the average occupancy column
        for col_num in range(1, len(row_data) + 2):
            cell = worksheet.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row+=1

    
    if record_state:
        worksheet[f'F{merge_start}']=round(occupancy/(merge_end-merge_start+1),2)
        occupancy=0
        worksheet.merge_cells(f'A{merge_start}:A{merge_end}')
        worksheet.merge_cells(f'F{merge_start}:F{merge_end}')
        thin=Side(border_style="thin",color="000000")
        border=Border(top=thin,left=thin,right=thin,bottom=thin)
        for row in worksheet[f'A3:F{merge_end}']:
            for cell in row:
                cell.border=border
        


    workbook.save(output)
    output.seek(0)

    exp_datestart=datetime.strptime(str_date_start, '%Y-%m-%d').strftime('%d %B %Y')
    exp_dateend=datetime.strptime(str_date_end, '%Y-%m-%d').strftime('%d %B %Y')

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment;filename=Laporan Kepadatan {exp_datestart}-{exp_dateend}.xlsx"
        }
    )

@app.route('/download_all',methods=['GET', 'POST'])
@login_req
def download_all():
    if request.method == 'POST':
        #gather all the data from database
        data = Occupancy.query.all()

    output=io.BytesIO()
    #create workbook object
    workbook=Workbook()
    # Initial Workbook Configuration - Sheet Name, Column Width, Merge Cell, Font Style
    worksheet=workbook.active
    worksheet.title="Laporan Jumlah Kepadatan Area"
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 7
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 20
    
    worksheet.merge_cells('A1:H1')
    font_title=Font(name='Calibri',size=14,bold=True)
    title_style=worksheet['A1']
    title_style.font=font_title

    worksheet['A1'] = f"Laporan Seluruh Data Jumlah Kepadatan Area"
    worksheet['A2']='' #break
    # Header Configuration
    header_contents=['Tanggal','Jam','Jumlah Masuk','Jumlah Keluar','Kepadatan Ruangan','Rerata Kepadatan']
    worksheet.append(header_contents)
    for col_num in range(1, len(header_contents) + 1):
        font_header=Font(name='Calibri',size=12,bold=True)
        cell = worksheet.cell(row=3, column=col_num)
        cell.font=font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
   
    # initiation of merging cells logic
    previous_date=None
    merge_start=4
    merge_end=0
    current_row=4
    record_state=False
    occupancy=0
    # iterate every row found
    for row in data:
        #if there's no previous date, set the previous date to the current date
        if previous_date==None:
            previous_date=row.createdAt.strftime("%d %B %Y")
            occupancy+=row.in_room
        #to check if the current date is the same as the previous date and the record state is false
        elif previous_date==row.createdAt.strftime("%d %B %Y") and record_state==False:
            #set the merge start and end to the current row
            merge_start=current_row-1
            merge_end=current_row
            record_state=True
            occupancy+=row.in_room
        # other branch: if the current date is the same as the previous date and the record state is true    
        elif previous_date==row.createdAt.strftime("%d %B %Y") and record_state==True:
            #will continue to set the merge end to the current row
            merge_end=current_row
            occupancy+=row.in_room
        # to represent that the date has changed and the record state is true
        elif previous_date!=row.createdAt.strftime("%d %B %Y") and record_state==True:
            # will merge the cells from the merge start to the merge end also set the record state to false and set the previous date to the current date
            worksheet.merge_cells(f'A{merge_start}:A{merge_end}')
            worksheet.merge_cells(f'F{merge_start}:F{merge_end}')
            # for row_num in range(merge_start, merge_end + 1):
            #     cell = worksheet.cell(row=row_num, column=1)
            #     cell.alignment = Alignment(wrap_text=True)
            record_state=False
            previous_date=row.createdAt.strftime("%d %B %Y")
            worksheet[f'F{merge_start}']=round(occupancy/(merge_end-merge_start+1),2)
            occupancy=0
        # Add New Row Data    
        day=format_date(row.createdAt, format='EEEE', locale=locale)
        date=row.createdAt.strftime("%d %B %Y")
        day_date=f"{day}, {date}"
        row_data=[day_date,row.createdAt.strftime("%H:%M"),row.enter_total,row.out_total,row.in_room]
        worksheet.append(row_data)

        # Apply center alignment to each cell in the current row
        # len(row_data) + 2 -> to include the average occupancy column
        for col_num in range(1, len(row_data) + 2):
            cell = worksheet.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row+=1

    
    if record_state:
        worksheet[f'F{merge_start}']=round(occupancy/(merge_end-merge_start+1),2)
        occupancy=0
        worksheet.merge_cells(f'A{merge_start}:A{merge_end}')
        worksheet.merge_cells(f'F{merge_start}:F{merge_end}')
        thin=Side(border_style="thin",color="000000")
        border=Border(top=thin,left=thin,right=thin,bottom=thin)
        for row in worksheet[f'A3:F{merge_end}']:
            for cell in row:
                cell.border=border
        


    workbook.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment;filename=Laporan Kepadatan Area - Seluruh Data.xlsx"
        }
    )

@app.route('/settings')
@login_req
def settings():
    settings_coordinates=getCoordinates()
    nama = session.get('username')
    # will render the index.html file present in templates folder
    return render_template('settings.html', data=settings_coordinates, nama=nama)

@app.route('/analytics')
@login_req
def analytics():
    nama = session.get('username')
    # will render the index.html file present in templates folder
    new_data = Occupancy(len(enter_list), len(out_list), len(enter_list)-len(out_list), func.now())
    return render_template('analytics.html', jumlah=str(len(people_list)), masuk=str(len(enter_list)), keluar=str(len(out_list)), nama=nama)

@app.route('/video_feed')
@login_req
def video_feed():
    return Response(annotatedStream(), mimetype='multipart/x-mixed-replace; boundary=frame')
# # socketio for listening disconnected client
# @socketio.on('disconnect')
# def destroy():
#     cap.release()
#     cv2.destroyAllWindows()
#     cursor.close()
#     mydb.close()

@app.route('/user')
@login_req
def user():
    nama = session.get('username')
    if session.get('id') != 1:
        return redirect(url_for('analytics'))
    data = db.session.query(User).all()[1:]
    return render_template('user.html', data=data, nama=nama)

@app.route('/adduser', methods=['GET','POST'])
@login_req
def adduser():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        db.session.add(User(username,password))
        db.session.commit()

        return redirect(url_for('user'))
    
@app.route('/deleteuser/<id>', methods=['POST', 'GET'])
@login_req
def deleteuser(id):
    data = User.query.filter_by(id=id).first()

    db.session.delete(data)
    db.session.commit()

    return redirect(url_for('user'))

@app.route('/edituser/<id>', methods=['POST','GET'])
@login_req
def edituser(id):
    data = User.query.filter_by(id=id).first()
    if request.method == 'POST':
        data.username = request.form['edit_username']
        if data.password != '':
            data.password = bcrypt.generate_password_hash(request.form['edit_password']).decode('UTF-8')

        db.session.add(data)
        db.session.commit()

        return redirect(url_for('user'))
# Buat database
db.create_all()
# Cek apakah admin user sudah ada dalam database
admin = User.query.filter_by(username='admin').first()
if not admin:
    admin = User('admin','admin')
    db.session.add(admin)
    db.session.commit()

# ========== TASK SCHEDULER RELATED FUNCTION (START) ===========
# # Initialize the APScheduler
# scheduler = BackgroundScheduler()
# # Repeat your dataInsertion function every 1 hour
# scheduler.add_job(submitData, 'interval', seconds=3600)
# # Start the scheduler when the Flask app starts
# scheduler.start()

scheduler = BackgroundScheduler()
scheduler.add_job(
    submitData,
    'cron',
    hour='8-18',  # Menjadwalkan mulai dari jam 08:00 sampai jam 17:00
    minute=0,  # Menit ke-0
)
scheduler.start()
    
if __name__ == '__main__':
    app.run(debug=True, port=40047, use_reloader=False)
